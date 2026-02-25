"""
Project class -- Layer 0 declarative API for PySchemaElectrical.

The Project is the top-level object that owns state, terminal registry,
circuit definitions, page layout, and output configuration. Users interact
with it declaratively to define an entire schematic drawing set and compile
it to a multi-page PDF.
"""

import os
import shutil
from collections.abc import Callable
from dataclasses import dataclass, field, replace
from typing import TYPE_CHECKING, Any

from pyschemaelectrical.builder import BuildResult, CircuitBuilder

if TYPE_CHECKING:
    from pyschemaelectrical.field_devices import ConnectionRow
    from pyschemaelectrical.plc_resolver import PlcRack

from pyschemaelectrical.descriptors import Descriptor, build_from_descriptors
from pyschemaelectrical.system.connection_registry import (
    export_registry_to_csv,
    get_registry,
)
from pyschemaelectrical.system.system import render_system
from pyschemaelectrical.terminal import Terminal
from pyschemaelectrical.utils.autonumbering import create_autonumberer
from pyschemaelectrical.utils.export_utils import (
    export_terminal_list,
    finalize_terminal_csv,
)

# ---------------------------------------------------------------------------
# Internal data structures
# ---------------------------------------------------------------------------


@dataclass
class _CircuitDef:
    """Internal deferred circuit definition."""

    key: str
    factory: str  # "descriptors" or "custom"
    params: dict[str, Any] = field(default_factory=dict)
    count: int = 1
    wire_labels: list[str] | None = None
    reuse_tags: dict[str, str] | None = None  # maps prefix -> circuit key
    components: list[Descriptor] | None = None
    builder_fn: Callable | None = None
    start_indices: dict[str, int] | None = None
    terminal_start_indices: dict[str, int] | None = None


@dataclass
class _PageDef:
    """Internal page definition."""

    page_type: str  # "schematic", "front", "terminal_report", "plc_report", "custom"
    title: str = ""
    circuit_key: str = ""
    circuit_keys: list[str] | None = None
    md_path: str = ""
    notice: str | None = None
    csv_path: str = ""
    typst_content: str = ""


# ---------------------------------------------------------------------------
# Project class
# ---------------------------------------------------------------------------


class Project:
    """Declarative project builder for electrical schematic drawing sets.

    Project is one of the intentional mutable builder classes in the library.
    It accumulates terminal definitions, circuit registrations, and page
    layouts, then compiles everything to a multi-page PDF via ``.build()``.

    State is threaded automatically between circuits in registration order,
    so terminal pin numbers auto-increment correctly across the drawing set.

    Example::

        project = Project(
            title="My Schematics",
            drawing_number="DWG-001",
            author="Engineer",
            project="Project Name",
        )
        project.terminals(
            Terminal("X1", "Main Power"),
            Terminal("X3", "Fused 24V", bridge="all"),
            Terminal("X4", "Ground", bridge="all"),
        )
        project.add_circuit("motors", my_builder, count=3)
        project.page("Motor Circuits", "motors")
        project.terminal_report()
        project.build("output.pdf")

    Warning:
        Do not share Project instances across multiple build contexts.
        Each Project should be used for a single ``.build()`` call.

    Args:
        title: Drawing title (appears in title block).
        drawing_number: Drawing number (appears in title block).
        author: Author name.
        project: Project name.
        revision: Revision string (e.g. "00", "A1").
        logo: Path to logo image file (optional).
        font: Font family for all text output.
    """

    def __init__(
        self,
        title: str = "",
        drawing_number: str = "",
        author: str = "",
        project: str = "",
        revision: str = "00",
        logo: str | None = None,
        font: str = "Times New Roman",
    ):
        self.title = title
        self.drawing_number = drawing_number
        self.author = author
        self.project = project
        self.revision = revision
        self.logo = logo
        self.font = font

        self._state = create_autonumberer()
        self._terminals: dict[str, Terminal] = {}
        self._circuit_defs: list[_CircuitDef] = []
        self._pages: list[_PageDef] = []
        self._results: dict[str, BuildResult] = {}
        self._plc_rack: "PlcRack | None" = None
        self._external_connections: "list[ConnectionRow]" = []
        self._field_device_defs: list[tuple[list, dict | None]] = []
        self._wire_label_export: tuple[str, dict[str, str] | None] | None = None
        self._taglist_export: str | None = None
        self._bom_excel_export: str | None = None

    # ------------------------------------------------------------------
    # Terminal registration
    # ------------------------------------------------------------------

    def terminals(self, *terminals: Terminal):
        """Register terminal block definitions for this project.

        Terminals carry metadata (description, bridge info, reference flag)
        used for reports and auto-generation. Must be called before
        registering circuits that reference these terminals.

        Args:
            *terminals: One or more ``Terminal`` instances.
        """
        for t in terminals:
            self._terminals[str(t)] = t

    @property
    def _terminal_descriptions(self) -> dict[str, str]:
        return {tag: t.title for tag, t in self._terminals.items() if t.title}

    def set_pin_start(self, terminal_id: str, pin: int) -> None:
        """Seed the pin counter for a terminal so auto-allocation starts at *pin*.

        Also updates per-prefix counters for this terminal so that
        prefixed allocations respect the new floor.

        Args:
            terminal_id: Terminal tag (e.g. "X1").
            pin: Starting pin number (subsequent auto-allocations will
                begin at ``pin + 1``).
        """
        tag_key = str(terminal_id)

        new_counters = {**self._state.terminal_counters, tag_key: pin}

        prefix_counters = self._state.terminal_prefix_counters
        if tag_key in prefix_counters:
            new_tag_prefixes = prefix_counters[tag_key].copy()
            for p in new_tag_prefixes:
                new_tag_prefixes[p] = pin
            new_prefix_counters = {**prefix_counters, tag_key: new_tag_prefixes}
        else:
            new_prefix_counters = prefix_counters

        self._state = replace(
            self._state,
            terminal_counters=new_counters,
            terminal_prefix_counters=new_prefix_counters,
        )

    # ------------------------------------------------------------------
    # Circuit registration
    # ------------------------------------------------------------------

    def circuit(
        self,
        key: str,
        components: list[Descriptor],
        count: int = 1,
        wire_labels: list[str] | None = None,
        reuse_tags: dict[str, str] | None = None,
        start_indices: dict[str, int] | None = None,
        terminal_start_indices: dict[str, int] | None = None,
        **kwargs,
    ):
        """
        Register a custom inline circuit from descriptors.

        Args:
            key: Unique circuit identifier.
            components: List of ref(), comp(), term() descriptors.
            count: Number of instances.
            wire_labels: Wire label strings.
            reuse_tags: Maps tag prefix to source circuit key.
            start_indices: Override tag counters.
            terminal_start_indices: Override terminal pin counters.
        """
        self._circuit_defs.append(
            _CircuitDef(
                key=key,
                factory="descriptors",
                count=count,
                wire_labels=wire_labels,
                reuse_tags=reuse_tags,
                components=components,
                params=kwargs,
                start_indices=start_indices,
                terminal_start_indices=terminal_start_indices,
            )
        )

    def custom(self, key: str, builder_fn: Callable, count: int = 1, **kwargs):
        """
        Register a custom circuit built via a builder function.

        The function receives ``(state, **kwargs)`` and must return
        a ``BuildResult`` (or a tuple ``(state, circuit, used_terminals)``).
        """
        self._circuit_defs.append(
            _CircuitDef(
                key=key,
                factory="custom",
                count=count,
                builder_fn=builder_fn,
                params=kwargs,
            )
        )

    def reserve_pins(self, key: str, terminal: "Terminal", count: int) -> "Project":
        """Reserve terminal pins with a bridge group (e.g., emergency stop).

        Registers a deferred circuit that advances the pin counter by *count*
        and creates a bridge group spanning those pins.  The circuit itself
        is empty (no visual elements).

        Args:
            key: Unique circuit identifier for this reservation.
            terminal: Terminal to reserve pins on.
            count: Number of sequential pins to reserve.

        Returns:
            self (for method chaining).
        """

        def _reserve_fn(state, **_kwargs):
            from pyschemaelectrical.system.system import Circuit
            from pyschemaelectrical.utils.autonumbering import (
                get_terminal_counter,
                set_terminal_counter,
            )

            tag = str(terminal)
            start = get_terminal_counter(state, tag) + 1
            end = start + count - 1
            state = set_terminal_counter(state, terminal, end)
            return BuildResult(
                state=state,
                circuit=Circuit(),
                used_terminals=[],
                bridge_groups={tag: [(start, end)]},
            )

        self._circuit_defs.append(
            _CircuitDef(key=key, factory="custom", builder_fn=_reserve_fn)
        )
        return self

    # ------------------------------------------------------------------
    # CircuitBuilder-based circuit registration
    # ------------------------------------------------------------------

    def add_circuit(
        self,
        name: str,
        builder: CircuitBuilder,
        *,
        count: int = 1,
        reuse_tags: dict[str, str] | None = None,
        reuse_terminals: dict[str, str] | None = None,
        start_indices: dict[str, int] | None = None,
        terminal_start_indices: dict[str, int] | None = None,
        tag_generators: dict | None = None,
        terminal_maps: dict | None = None,
        wire_labels: list[str] | None = None,
    ) -> CircuitBuilder:
        """Build a CircuitBuilder immediately and register its result.

        Builds the circuit right away and stores the ``BuildResult``.
        State is advanced so subsequent circuits see the updated
        tag/terminal counters.

        Args:
            name: Unique circuit identifier (used as key in results and
                for SVG/CSV filenames).
            builder: An un-built (not frozen) ``CircuitBuilder`` instance.
            count: Number of circuit instances to create.
            reuse_tags: Dict mapping tag prefix to the *name* of a
                previously added circuit whose tags should be reused.
                E.g. ``{"K": "coils"}`` reuses K tags from the circuit
                registered under the name ``"coils"``.
            reuse_terminals: Dict mapping terminal key to the *name* of
                a previously added circuit whose terminal pins should be
                reused. E.g. ``{"X008": "pumps"}`` reuses X008 pins.
            start_indices: Override tag counters (e.g. ``{"K": 3}``).
            terminal_start_indices: Override terminal pin counters.
            tag_generators: Custom tag generator functions or fixed-tag
                strings. Merged with resolved reuse generators (manual
                overrides win).
            terminal_maps: Terminal ID overrides by logical name.
            wire_labels: Wire label strings for vertical wires.

        Returns:
            The frozen ``CircuitBuilder`` (callers can use it for inline
            reuse via ``builder.reuse_tags()`` etc.).

        Raises:
            RuntimeError: If the builder is already frozen.
            ValueError: If a reuse_tags/reuse_terminals reference names a
                circuit that hasn't been added yet.
        """
        if builder._frozen:
            raise RuntimeError(
                f"Cannot add_circuit('{name}'): the CircuitBuilder is already "
                f"frozen. Pass an un-built builder."
            )

        # Resolve reuse_tags: name -> BuildResult
        resolved_reuse_tags: (
            dict[str, BuildResult | CircuitBuilder | Callable] | None
        ) = None
        if reuse_tags:
            resolved_reuse_tags = {}
            for prefix, source_name in reuse_tags.items():
                if source_name not in self._results:
                    raise ValueError(
                        f"add_circuit('{name}') references '{source_name}' via "
                        f"reuse_tags['{prefix}'], but it hasn't been added yet. "
                        f"Add '{source_name}' before '{name}'."
                    )
                resolved_reuse_tags[prefix] = self._results[source_name]

        # Resolve reuse_terminals: name -> BuildResult
        resolved_reuse_terminals: (
            dict[str, BuildResult | CircuitBuilder | Callable] | None
        ) = None
        if reuse_terminals:
            resolved_reuse_terminals = {}
            for key, source_name in reuse_terminals.items():
                if source_name not in self._results:
                    raise ValueError(
                        f"add_circuit('{name}') references '{source_name}' via "
                        f"reuse_terminals['{key}'], but it hasn't been added yet. "
                        f"Add '{source_name}' before '{name}'."
                    )
                resolved_reuse_terminals[key] = self._results[source_name]

        builder.build(
            state=self._state,
            count=count,
            reuse_tags=resolved_reuse_tags,
            reuse_terminals=resolved_reuse_terminals,
            start_indices=start_indices,
            terminal_start_indices=terminal_start_indices,
            tag_generators=tag_generators,
            terminal_maps=terminal_maps,
            wire_labels=wire_labels,
        )

        result = builder._result
        assert result is not None  # guaranteed after successful build()
        self._results[name] = result
        self._state = builder.state
        return builder

    # ------------------------------------------------------------------
    # PLC rack and external connections
    # ------------------------------------------------------------------

    def plc_rack(self, rack: "PlcRack") -> "Project":
        """Register a PLC rack for automatic PLC connection report generation.

        When a rack is registered, ``.build()`` will automatically generate
        the PLC connections CSV from the circuit registry and any registered
        external connections.

        Args:
            rack: List of (designation, PlcModuleType) tuples describing
                the physical PLC rack.

        Returns:
            self (for method chaining).
        """
        self._plc_rack = rack
        return self

    def external_connections(self, connections: "list[ConnectionRow]") -> "Project":
        """Register external field wiring connections for the PLC report.

        These are connections from field devices (sensors, valves, motors)
        entering the cabinet. They are resolved against the PLC rack to
        generate the PLC connection report.

        Args:
            connections: List of ConnectionRow tuples
                (component_from, pin_from, terminal_tag, terminal_pin,
                 component_to, pin_to).

        Returns:
            self (for method chaining).
        """
        self._external_connections.extend(connections)
        return self

    def add_field_devices(
        self,
        connections: "list[ConnectionRow]",
        reuse_terminals: dict[str, str] | None = None,
    ) -> "Project":
        """Register external field device connections.

        Alias for ``external_connections()`` with an optional
        ``reuse_terminals`` parameter for future expansion.

        Args:
            connections: List of ``ConnectionRow`` tuples.
            reuse_terminals: Reserved for future use. Currently ignored.

        Returns:
            self (for method chaining).
        """
        self._external_connections.extend(connections)
        return self

    def field_devices(
        self,
        devices: list,
        reuse_terminals: dict | None = None,
        template_reuse: dict | None = None,
    ) -> "Project":
        """Register field devices for deferred connection resolution.

        After build_circuits(), resolves reuse_terminals from built circuit
        results, generates connections via generate_field_connections(),
        and resolves PLC references if a rack is registered.

        Args:
            devices: List of FieldDevice instances.
            reuse_terminals: Maps Terminal -> circuit key string.
                Pins from that circuit's terminal_pin_map are reused.
            template_reuse: Maps DeviceTemplate -> {Terminal: circuit key}.
                Only devices whose template matches will reuse those
                terminal pins; other devices auto-number normally but
                skip the reserved pin values.

        Returns:
            self (for method chaining).
        """
        self._field_device_defs.append((devices, reuse_terminals, template_reuse))
        return self

    # ------------------------------------------------------------------
    # Query properties
    # ------------------------------------------------------------------

    @property
    def device_registry(self) -> dict:
        """Merged device registry from all built circuits."""
        merged: dict = {}
        for result in self._results.values():
            merged.update(result.device_registry)
        return merged

    @property
    def bridge_groups(self) -> dict:
        """Merged bridge groups from all built circuits."""
        merged: dict = {}
        for result in self._results.values():
            for key, groups in result.bridge_groups.items():
                merged.setdefault(key, []).extend(groups)
        return merged

    @property
    def wire_connections(self) -> dict[str, list]:
        """Wire connections grouped by circuit name."""
        return {
            key: result.wire_connections
            for key, result in self._results.items()
            if result.wire_connections
        }

    @property
    def resolved_connections(self) -> list:
        """All resolved external connections (available after build_circuits())."""
        return list(self._external_connections)

    # ------------------------------------------------------------------
    # Page management
    # ------------------------------------------------------------------

    def page(self, title: str, circuit_key: str | list[str]):
        """Add a schematic page to the PDF output.

        Args:
            title: Page title displayed in the title block.
            circuit_key: Key of a registered circuit to render, or a list of
                keys to merge onto a single page.
        """
        if isinstance(circuit_key, list):
            self._pages.append(
                _PageDef(page_type="schematic", title=title, circuit_keys=circuit_key)
            )
        else:
            self._pages.append(
                _PageDef(page_type="schematic", title=title, circuit_key=circuit_key)
            )

    def front_page(self, md_path: str, notice: str | None = None):
        """Add a front page rendered from a Markdown file.

        Args:
            md_path: Path to the Markdown source file.
            notice: Optional notice text displayed on the front page.
        """
        self._pages.append(_PageDef(page_type="front", md_path=md_path, notice=notice))

    def terminal_report(self):
        """Add an auto-generated system terminal report page.

        Includes all registered terminals with bridge/connection info
        and descriptions from ``terminals()``.
        """
        self._pages.append(_PageDef(page_type="terminal_report"))

    def plc_report(self, csv_path: str = "") -> "Project":
        """Add a PLC connections report page.

        When a rack has been registered via ``plc_rack()`` and *csv_path*
        is empty, ``.build()`` will auto-generate the PLC connections CSV
        from the circuit registry and any registered external connections.

        Args:
            csv_path: Path to the PLC connections CSV file.  Leave empty
                when using the auto-generation path via ``plc_rack()``.

        Returns:
            self (for method chaining).
        """
        self._pages.append(_PageDef(page_type="plc_report", csv_path=csv_path))
        return self

    def custom_page(self, title: str, typst_content: str):
        """Add a page with raw Typst markup content.

        Args:
            title: Page title displayed in the title block.
            typst_content: Raw Typst source code for the page body.
        """
        self._pages.append(
            _PageDef(page_type="custom", title=title, typst_content=typst_content)
        )

    def bom_report(self) -> "Project":
        """Add an auto-generated Bill of Materials page."""
        self._pages.append(_PageDef(page_type="bom_report"))
        return self

    def export_wire_labels(
        self, path: str, titles: dict[str, str] | None = None
    ) -> "Project":
        """Register a wire label CSV export. Written during build()."""
        self._wire_label_export = (path, titles)
        return self

    def export_taglist(self, path: str) -> "Project":
        """Register a taglist CSV export. Written during build()."""
        self._taglist_export = path
        return self

    def export_bom_excel(self, path: str) -> "Project":
        """Register a BOM Excel export. Written during build()."""
        self._bom_excel_export = path
        return self

    def export_cable_csv(
        self, output_path: str
    ) -> "tuple[str, int, dict[str, str], dict[str, dict]]":
        """Generate wireviz-compatible cable CSV from field devices.

        Requires build_circuits() to have been called first.

        Args:
            output_path: Path for the output CSV file.

        Returns:
            (csv_path, cable_count, cable_titles, connector_overrides)
        """
        from pyschemaelectrical.cable_export import generate_cable_csv

        all_devices = []
        for devices, _reuse in self._field_device_defs:
            all_devices.extend(devices)

        return generate_cable_csv(self._external_connections, all_devices, output_path)

    # ------------------------------------------------------------------
    # Build pipeline
    # ------------------------------------------------------------------

    def build_circuits(self) -> None:
        """Build all deferred circuits and resolve field devices."""
        self._build_all_circuits()
        self._resolve_field_devices()

    def build(
        self,
        output: str,
        temp_dir: str = "temp",
        keep_temp: bool = False,
    ):
        """
        Build all circuits and compile the PDF.

        Steps:
        1. Build all registered circuits (respecting dependencies).
        2. Generate SVG for each circuit.
        3. Generate per-circuit terminal CSV.
        4. Generate system terminal CSV with bridge info.
        5. Assemble and compile Typst -> PDF.

        Args:
            output: Path for the output PDF file.
            temp_dir: Directory for intermediate files.
            keep_temp: If True, keep intermediate files after compilation.
        """
        from pyschemaelectrical.rendering.typst.compiler import (
            TypstCompiler,
            TypstCompilerConfig,
        )

        os.makedirs(temp_dir, exist_ok=True)

        # 1. Build all circuits
        self._build_all_circuits()
        self._resolve_field_devices()

        # 2. Generate SVGs and terminal CSVs
        svg_paths = {}
        csv_paths = {}

        for key, result in self._results.items():
            svg_path = os.path.join(temp_dir, f"{key}.svg")
            render_system(result.circuit, svg_path)
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = os.path.join(temp_dir, f"{key}_terminals.csv")
                export_terminal_list(csv_path, result.used_terminals, self._terminal_descriptions)
                csv_paths[key] = csv_path

        self._render_multi_circuit_pages(svg_paths, csv_paths, temp_dir)
        self._export_wire_labels()
        self._export_taglist()
        self._export_bom_excel()

        # 3. Generate system terminal CSV with bridge info
        system_csv_path = self._generate_system_csv(temp_dir)

        # 3.5. Auto-generate PLC connections CSV if rack is configured
        plc_csv_path = ""
        if self._plc_rack is not None:
            plc_csv_path = os.path.join(temp_dir, "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

        # 4. Assemble Typst document
        # Use CWD as root so all relative paths (SVGs, CSVs) resolve correctly
        root_dir = os.getcwd()
        config = TypstCompilerConfig(
            drawing_name=self.title,
            drawing_number=self.drawing_number,
            author=self.author,
            project=self.project,
            revision=self.revision,
            logo_path=os.path.abspath(self.logo) if self.logo else None,
            font_family=self.font,
            root_dir=root_dir,
            temp_dir=os.path.relpath(os.path.abspath(temp_dir), root_dir),
        )
        compiler = TypstCompiler(config)

        # Add pages
        for page_def in self._pages:
            self._add_page_to_compiler(
                compiler, page_def, svg_paths, csv_paths, system_csv_path, plc_csv_path
            )

        # 5. Compile
        compiler.compile(output)
        print(f"PDF compiled: {output}")

        # Cleanup
        if not keep_temp:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Build SVGs only (no PDF, no typst dependency)
    # ------------------------------------------------------------------

    def build_svgs(self, output_dir: str = "output"):
        """
        Build all circuits and export SVGs (no PDF compilation).

        Useful when the ``typst`` package is not installed.

        Args:
            output_dir: Directory for output SVG and CSV files.
        """
        os.makedirs(output_dir, exist_ok=True)
        self._build_all_circuits()

        svg_paths: dict[str, str] = {}
        csv_paths: dict[str, str] = {}

        for key, result in self._results.items():
            svg_path = os.path.join(output_dir, f"{key}.svg")
            render_system(result.circuit, svg_path)
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = os.path.join(output_dir, f"{key}_terminals.csv")
                export_terminal_list(csv_path, result.used_terminals, self._terminal_descriptions)
                csv_paths[key] = csv_path

        self._render_multi_circuit_pages(svg_paths, csv_paths, output_dir)
        self._export_wire_labels()
        self._export_taglist()

        # System terminal CSV
        self._generate_system_csv(output_dir)

        print(f"SVGs and CSVs written to: {output_dir}")

    # ------------------------------------------------------------------
    # Separate output methods
    # ------------------------------------------------------------------

    def render_svgs(self, output_dir: str) -> None:
        """Render all circuit SVGs and per-circuit terminal CSVs to *output_dir*.

        Unlike ``build_svgs()``, this method does **not** build deferred
        circuits. It only renders results already present in ``_results``
        (populated by ``add_circuit()`` or a prior ``_build_all_circuits()``
        call).

        Args:
            output_dir: Directory for output SVG and CSV files.
        """
        os.makedirs(output_dir, exist_ok=True)

        for key, result in self._results.items():
            svg_path = os.path.join(output_dir, f"{key}.svg")
            render_system(result.circuit, svg_path)

            if result.used_terminals:
                csv_path = os.path.join(output_dir, f"{key}_terminals.csv")
                export_terminal_list(csv_path, result.used_terminals, self._terminal_descriptions)

    def export_csvs(self, output_dir: str) -> None:
        """Export system terminal CSV with bridge info to *output_dir*.

        Also generates the PLC connections CSV when a rack has been
        registered via ``plc_rack()``.

        Args:
            output_dir: Directory for output CSV files.
        """
        os.makedirs(output_dir, exist_ok=True)

        # System terminal CSV
        self._generate_system_csv(output_dir)

        # PLC connections CSV
        if self._plc_rack is not None:
            plc_csv_path = os.path.join(output_dir, "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

    def compile_pdf(
        self,
        output: str,
        temp_dir: str = "temp",
        keep_temp: bool = False,
    ) -> None:
        """Compile the full PDF using TypstCompiler with the defined page flow.

        This method renders SVGs, exports CSVs, and compiles the Typst
        document into a single PDF. It operates on results already present
        in ``_results`` (populated by ``add_circuit()`` or a prior
        ``_build_all_circuits()`` call).

        Args:
            output: Path for the output PDF file.
            temp_dir: Directory for intermediate files.
            keep_temp: If True, keep intermediate files after compilation.
        """
        from pyschemaelectrical.rendering.typst.compiler import (
            TypstCompiler as _TypstCompiler,
        )
        from pyschemaelectrical.rendering.typst.compiler import (
            TypstCompilerConfig,
        )

        os.makedirs(temp_dir, exist_ok=True)

        # Render SVGs and per-circuit terminal CSVs
        svg_paths: dict[str, str] = {}
        csv_paths: dict[str, str] = {}

        for key, result in self._results.items():
            svg_path = os.path.join(temp_dir, f"{key}.svg")
            render_system(result.circuit, svg_path)
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = os.path.join(temp_dir, f"{key}_terminals.csv")
                export_terminal_list(csv_path, result.used_terminals, self._terminal_descriptions)
                csv_paths[key] = csv_path

        self._render_multi_circuit_pages(svg_paths, csv_paths, temp_dir)
        self._export_wire_labels()
        self._export_taglist()

        # System terminal CSV with bridge info
        system_csv_path = self._generate_system_csv(temp_dir)

        # PLC connections CSV
        plc_csv_path = ""
        if self._plc_rack is not None:
            plc_csv_path = os.path.join(temp_dir, "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

        # Assemble Typst document
        root_dir = os.getcwd()
        config = TypstCompilerConfig(
            drawing_name=self.title,
            drawing_number=self.drawing_number,
            author=self.author,
            project=self.project,
            revision=self.revision,
            logo_path=os.path.abspath(self.logo) if self.logo else None,
            font_family=self.font,
            root_dir=root_dir,
            temp_dir=os.path.relpath(os.path.abspath(temp_dir), root_dir),
        )
        compiler = _TypstCompiler(config)

        for page_def in self._pages:
            self._add_page_to_compiler(
                compiler, page_def, svg_paths, csv_paths, system_csv_path, plc_csv_path
            )

        compiler.compile(output)
        print(f"PDF compiled: {output}")

        if not keep_temp:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Internal: field device resolution
    # ------------------------------------------------------------------

    def _resolve_field_devices(self) -> None:
        """Resolve deferred field device registrations."""
        if not self._field_device_defs:
            return

        from pyschemaelectrical.field_devices import generate_field_connections

        for devices, reuse_terminals, template_reuse in self._field_device_defs:
            resolved_reuse = None
            if reuse_terminals:
                resolved_reuse = {}
                for terminal, circuit_key in reuse_terminals.items():
                    if circuit_key not in self._results:
                        raise ValueError(
                            f"field_devices() references circuit '{circuit_key}' "
                            f"for terminal reuse, but it hasn't been built yet."
                        )
                    resolved_reuse[str(terminal)] = self._results[circuit_key]

            resolved_template_reuse = None
            if template_reuse:
                resolved_template_reuse = {}
                for tmpl, terminal_map in template_reuse.items():
                    resolved_template_reuse[tmpl] = {}
                    for terminal, circuit_key in terminal_map.items():
                        if circuit_key not in self._results:
                            raise ValueError(
                                f"field_devices() template_reuse references "
                                f"circuit '{circuit_key}', but it hasn't "
                                f"been built yet."
                            )
                        resolved_template_reuse[tmpl][str(terminal)] = (
                            self._results[circuit_key]
                        )

            connections = generate_field_connections(
                devices,
                reuse_terminals=resolved_reuse,
                template_reuse=resolved_template_reuse,
            )

            if self._plc_rack:
                from pyschemaelectrical.plc_resolver import resolve_plc_references

                connections = resolve_plc_references(connections, self._plc_rack)

            self._external_connections.extend(connections)

    # ------------------------------------------------------------------
    # Internal: circuit building
    # ------------------------------------------------------------------

    def _build_all_circuits(self):
        """Build all registered circuits in order."""
        self._results = {}
        for cdef in self._circuit_defs:
            result = self._build_one_circuit(cdef)
            self._results[cdef.key] = result
            self._state = result.state

    def _build_one_circuit(self, cdef: _CircuitDef) -> BuildResult:
        """Build a single circuit definition."""
        # Resolve reuse_tags: map circuit key -> BuildResult
        resolved_reuse = None
        if cdef.reuse_tags:
            resolved_reuse = {}
            for prefix, source_key in cdef.reuse_tags.items():
                if source_key not in self._results:
                    raise ValueError(
                        f"Circuit '{cdef.key}' references '{source_key}' via "
                        f"reuse_tags, but it hasn't been built yet. "
                        f"Register '{source_key}' before '{cdef.key}'."
                    )
                resolved_reuse[prefix] = self._results[source_key]

        if cdef.factory == "descriptors":
            return self._build_descriptor_circuit(cdef, resolved_reuse)
        elif cdef.factory == "custom":
            return self._build_custom_circuit(cdef)
        else:
            raise ValueError(
                f"Unknown circuit factory '{cdef.factory}'. "
                f"Use 'descriptors' or 'custom'."
            )

    def _build_descriptor_circuit(
        self, cdef: _CircuitDef, resolved_reuse: dict | None
    ) -> BuildResult:
        """Build a circuit from inline descriptors."""
        if cdef.components is None:
            raise ValueError(
                f"Circuit '{cdef.key}' uses descriptor mode but has no "
                f"components defined"
            )
        return build_from_descriptors(
            self._state,
            cdef.components,
            x=cdef.params.get("x", 0.0),
            y=cdef.params.get("y", 0.0),
            spacing=cdef.params.get("spacing", 80.0),
            count=cdef.count,
            wire_labels=cdef.wire_labels,
            reuse_tags=resolved_reuse,
            start_indices=cdef.start_indices,
            terminal_start_indices=cdef.terminal_start_indices,
        )

    def _build_custom_circuit(self, cdef: _CircuitDef) -> BuildResult:
        """Build a circuit via user-provided builder function."""
        if cdef.builder_fn is None:
            raise ValueError(
                f"Circuit '{cdef.key}' uses custom mode but has no builder_fn defined"
            )
        result = cdef.builder_fn(self._state, **cdef.params)
        if isinstance(result, BuildResult):
            return result
        # Support frozen CircuitBuilder return
        if isinstance(result, CircuitBuilder) and result._frozen:
            return result._result
        # Support tuple return: (state, circuit, used_terminals)
        state, circuit, used_terminals = result
        return BuildResult(
            state=state,
            circuit=circuit,
            used_terminals=used_terminals,
        )

    def _render_multi_circuit_pages(self, svg_paths, csv_paths, output_dir):
        """Render merged SVGs for multi-circuit pages."""
        from pyschemaelectrical.builder import merge_build_results

        for page_def in self._pages:
            if page_def.circuit_keys:
                results_to_merge = [
                    self._results[k]
                    for k in page_def.circuit_keys
                    if k in self._results
                ]
                if results_to_merge:
                    merged = merge_build_results(results_to_merge)
                    merged_key = "_".join(page_def.circuit_keys)
                    svg_path = os.path.join(output_dir, f"{merged_key}.svg")
                    render_system(merged.circuit, svg_path)
                    svg_paths[merged_key] = svg_path
                    if merged.used_terminals:
                        csv_path_m = os.path.join(
                            output_dir, f"{merged_key}_terminals.csv"
                        )
                        export_terminal_list(csv_path_m, merged.used_terminals, self._terminal_descriptions)
                        csv_paths[merged_key] = csv_path_m
                    # Point page_def to merged key for compiler
                    page_def.circuit_key = merged_key

    # ------------------------------------------------------------------
    # Internal: wire label and taglist exports
    # ------------------------------------------------------------------

    def _export_wire_labels(self) -> None:
        if self._wire_label_export is None:
            return
        import csv as _csv

        path, titles = self._wire_label_export
        titles = titles or {}
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        with open(path, "w", newline="") as f:
            writer = _csv.writer(f)
            for circuit_key, result in self._results.items():
                if not result.wire_connections:
                    continue
                title = titles.get(circuit_key, circuit_key)
                writer.writerow([title])
                for tag_a, pin_a, tag_b, pin_b in result.wire_connections:
                    writer.writerow([f"{tag_a}:{pin_a}"])
                    writer.writerow([f"{tag_b}:{pin_b}"])
                writer.writerow([])

    def _export_taglist(self) -> None:
        if self._taglist_export is None:
            return
        import csv as _csv

        from pyschemaelectrical.utils.utils import natural_sort_key

        tags: set[str] = set()
        for result in self._results.values():
            tags.update(result.device_registry.keys())
        for tid in self._terminals:
            tags.add(tid)
        if self._plc_rack:
            for slot_name, _module in self._plc_rack:
                tags.add(slot_name)

        os.makedirs(
            os.path.dirname(os.path.abspath(self._taglist_export)), exist_ok=True
        )
        with open(self._taglist_export, "w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(["Tag"])
            for tag in sorted(tags, key=natural_sort_key):
                writer.writerow([tag])

    def _export_bom_excel(self) -> None:
        if self._bom_excel_export is None:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        rows = self._aggregate_bom()
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        headers = ["Tags", "MPN", "Description", "Qty"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        for row_idx, (tags, mpn, desc, qty) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=tags)
            ws.cell(row=row_idx, column=2, value=mpn)
            ws.cell(row=row_idx, column=3, value=desc)
            ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 8

        path = self._bom_excel_export
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb.save(path)

    # ------------------------------------------------------------------
    # Internal: system CSV generation
    # ------------------------------------------------------------------

    def _generate_system_csv(self, output_dir: str) -> str:
        """Generate system terminal CSV with bridge info and external connections.

        PLC-prefixed connections are filtered out (they appear in the
        PLC report instead).
        """
        from pyschemaelectrical.system.connection_registry import TerminalRegistry

        csv_path = os.path.join(output_dir, "system_terminals.csv")
        registry = get_registry(self._state)
        filtered = tuple(
            c for c in registry.connections if not c.terminal_tag.startswith("PLC:")
        )
        registry = TerminalRegistry(connections=filtered)
        export_registry_to_csv(registry, csv_path, state=self._state)

        # Bridge defs from Terminal objects
        bridge_defs: dict = {}
        prefix_bridge_tags: set[str] = set()
        for tid, t in self._terminals.items():
            if t.bridge and not t.reference:
                if t.bridge == "per_prefix":
                    prefix_bridge_tags.add(tid)
                else:
                    bridge_defs[tid] = t.bridge

        # Bridge groups from circuit results
        for result in self._results.values():
            for key, groups in result.bridge_groups.items():
                bridge_defs.setdefault(key, []).extend(groups)

        finalize_terminal_csv(
            csv_path,
            bridge_defs=bridge_defs or None,
            prefix_bridge_tags=prefix_bridge_tags or None,
            external_connections=self._external_connections or None,
        )
        return csv_path

    # ------------------------------------------------------------------
    # Internal: BOM aggregation
    # ------------------------------------------------------------------

    def _count_terminal_pins(self) -> dict[str, int]:
        """Count unique pins per terminal from all connection sources."""
        pin_sets: dict[str, set] = {}
        for conn in get_registry(self._state).connections:
            pin_sets.setdefault(conn.terminal_tag, set()).add(conn.terminal_pin)
        for row in self._external_connections:
            tag, pin = str(row[2]), row[3]
            if tag and pin:
                pin_sets.setdefault(tag, set()).add(pin)
        return {k: len(v) for k, v in pin_sets.items()}

    def _aggregate_bom(self) -> list[tuple[str, str, str, int]]:
        """Aggregate BOM from device registries, terminals, and PLC modules."""
        from pyschemaelectrical.utils.utils import natural_sort_key

        terminal_pin_counts = self._count_terminal_pins()

        # Devices
        device_groups: dict[tuple[str, str], list[str]] = {}
        for result in self._results.values():
            for tag, device in result.device_registry.items():
                key = (device.mpn, device.description)
                device_groups.setdefault(key, []).append(tag)

        rows: list[tuple[str, str, str, int]] = []
        for (mpn, desc), tags in sorted(device_groups.items()):
            unique_tags = sorted(set(tags), key=natural_sort_key)
            rows.append(("/".join(unique_tags), mpn, desc, len(unique_tags)))

        # Terminals
        terminal_groups: dict[tuple[str, str], list[tuple[str, int]]] = {}
        for tid, t in self._terminals.items():
            if not t.reference and t.mpn:
                key = (t.mpn, t.description)
                pin_count = terminal_pin_counts.get(tid, 0)
                terminal_groups.setdefault(key, []).append((tid, pin_count))

        for (mpn, desc), entries in sorted(terminal_groups.items()):
            tags = sorted([e[0] for e in entries], key=natural_sort_key)
            total_pins = sum(e[1] for e in entries)
            rows.append(("/".join(tags), mpn, desc, total_pins))

        # PLC modules
        if self._plc_rack:
            plc_groups: dict[str, list[str]] = {}
            plc_desc: dict[str, str] = {}
            for slot_name, module in self._plc_rack:
                plc_groups.setdefault(module.mpn, []).append(slot_name)
                plc_desc[module.mpn] = (
                    f"PLC {module.signal_type} module ({module.channels}ch)"
                )
            for mpn, slots in sorted(plc_groups.items()):
                sorted_slots = sorted(slots, key=natural_sort_key)
                rows.append(
                    ("/".join(sorted_slots), mpn, plc_desc[mpn], len(sorted_slots))
                )

        return rows

    def _generate_bom_typst(self, bom_rows: list[tuple[str, str, str, int]]) -> str:
        """Generate Typst markup for BOM table."""
        lines = [
            "#place(bottom + center, dy: -title_offset)[",
            '  #text(size: 18pt, weight: "bold")[Bill of Materials]',
            "]",
            "#pad(left: 25mm, right: 25mm, top: 40mm, bottom: 40mm)[",
            "  #columns(2, gutter: 30em)[",
            "    #block(breakable: true)[",
            "      #table(",
            "        columns: (4.5cm, 3.5cm, 1fr, 1cm),",
            "        align: (left, left, left, right),",
            "        fill: (x, y) => if y == 0 { gray.lighten(85%) } else { none },",
            "        inset: 4pt,",
            "        stroke: 0.25pt + gray,",
            "        table.header(",
            '          text(size: 9pt, weight: "bold")[Tags],',
            '          text(size: 9pt, weight: "bold")[MPN],',
            '          text(size: 9pt, weight: "bold")[Description],',
            '          text(size: 9pt, weight: "bold")[Qty],',
            "        ),",
        ]
        for tags, mpn, desc, qty in bom_rows:
            tags_esc = tags.replace("#", "\\#")
            mpn_esc = mpn.replace("#", "\\#")
            desc_esc = desc.replace("#", "\\#")
            lines.append(
                f"        text(size: 9pt)[{tags_esc}], "
                f"text(size: 9pt)[{mpn_esc}], "
                f"text(size: 9pt)[{desc_esc}], "
                f"text(size: 9pt)[{qty}],"
            )
        lines.append("      )")
        lines.append("    ]")
        lines.append("  ]")
        lines.append("]")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Internal: page compilation
    # ------------------------------------------------------------------

    def _add_page_to_compiler(
        self,
        compiler: Any,
        page_def: _PageDef,
        svg_paths: dict[str, str],
        csv_paths: dict[str, str],
        system_csv_path: str,
        plc_csv_path: str = "",
    ):
        """Add a page definition to the TypstCompiler."""
        if page_def.page_type == "schematic":
            key = page_def.circuit_key
            svg_path = svg_paths.get(key, "")
            csv_path = csv_paths.get(key)
            if svg_path:
                compiler.add_schematic_page(page_def.title, svg_path, csv_path)
        elif page_def.page_type == "front":
            compiler.add_front_page(page_def.md_path, notice=page_def.notice)
        elif page_def.page_type == "terminal_report":
            titles = {
                str(t): t.title for t in self._terminals.values() if not t.reference
            }
            compiler.add_terminal_report(system_csv_path, titles)
        elif page_def.page_type == "plc_report":
            csv_path = page_def.csv_path or plc_csv_path
            if csv_path:
                compiler.add_plc_report(csv_path)
        elif page_def.page_type == "custom":
            compiler.add_custom_page(page_def.title, page_def.typst_content)
        elif page_def.page_type == "bom_report":
            bom_rows = self._aggregate_bom()
            typst_content = self._generate_bom_typst(bom_rows)
            compiler.add_custom_page("Bill of Materials", typst_content)

    # ------------------------------------------------------------------
    # Internal: PLC CSV generation
    # ------------------------------------------------------------------

    def _generate_plc_csv(self, csv_path: str) -> None:
        """Generate PLC connections CSV from registry and external connections."""
        import csv as _csv

        from pyschemaelectrical.plc_resolver import (
            extract_plc_connections_from_registry,
            generate_plc_report_rows,
            resolve_plc_references,
        )

        # _generate_plc_csv is only called when _plc_rack is not None
        rack = self._plc_rack
        assert rack is not None

        # Resolve external connections if any
        external = list(self._external_connections)
        if external:
            external = resolve_plc_references(external, rack)

        # Extract registry connections
        registry_connections: list[ConnectionRow] = (
            extract_plc_connections_from_registry(self._state, rack, external)
        )

        # Merge and generate rows
        all_connections = external + registry_connections
        rows = generate_plc_report_rows(all_connections, rack)

        with open(csv_path, "w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(
                ["Module", "MPN", "PLC Pin", "Component", "Pin", "Terminal"]
            )
            writer.writerows(rows)
